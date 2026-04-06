from __future__ import annotations

import math
import re
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from app.config import AppConfig, BusinessConfig, StoreMapping

try:
    from pyxlsb import open_workbook
except ImportError:  # pragma: no cover - dependency check happens at runtime
    open_workbook = None


class ParsingError(ValueError):
    pass


@dataclass(frozen=True)
class DailySalesRow:
    source_name: str
    gross_sales: float
    discount_amount: float
    cash_sales: float
    electronic_money_sales: float


@dataclass(frozen=True)
class ParsedCardSales:
    account_date: str | None
    by_store: dict[str, dict[str, float]]


@dataclass(frozen=True)
class ParsedDailySales:
    account_date: str | None
    by_store: dict[str, DailySalesRow]


@dataclass(frozen=True)
class ParsedSettlementOrders:
    account_date: str | None
    by_store: dict[str, float]
    by_merchant: dict[str, float]
    matched_merchants: dict[str, str]
    unmatched_merchants: tuple[str, ...]


@dataclass(frozen=True)
class ParsedVoucherSettlement:
    account_date: str | None
    by_output_store: dict[str, float]


@dataclass(frozen=True)
class SalesFormatInputRow:
    account_date: str
    source_name: str
    partner_code: str
    card_amount: float
    cash_amount: float
    discount_amount: float


@dataclass(frozen=True)
class ParsedSalesFormatInput:
    rows: tuple[SalesFormatInputRow, ...]
    parsed_store_count: int
    skipped_store_names: tuple[str, ...]
    notes: tuple[str, ...]
    date_min: str | None
    date_max: str | None


def _normalize_date_digits(digits: str) -> str | None:
    normalized = digits
    if len(normalized) == 6:
        normalized = f"20{normalized}"
    if len(normalized) != 8:
        return None
    try:
        datetime.strptime(normalized, "%Y%m%d")
    except ValueError:
        return None
    return normalized


def _as_yyyymmdd(value: Any) -> str | None:
    if isinstance(value, datetime):
        return value.strftime("%Y%m%d")
    if isinstance(value, date):
        return value.strftime("%Y%m%d")
    if isinstance(value, str):
        digits = re.sub(r"[^0-9]", "", value)
        return _normalize_date_digits(digits)
    return None


def _extract_account_date_from_filename(filename: str) -> str | None:
    match_with_separator = re.search(r"(20\d{2})[-_.](\d{2})[-_.](\d{2})", filename)
    if match_with_separator:
        return _normalize_date_digits("".join(match_with_separator.groups()))

    match_8 = re.search(r"(20\d{6})", filename)
    if match_8:
        return _normalize_date_digits(match_8.group(1))

    match_6 = re.search(r"(?<!\d)(\d{6})(?!\d)", filename)
    if match_6:
        return _normalize_date_digits(match_6.group(1))

    return None


def _extract_header_row(worksheet, required_headers: set[str], search_rows: int = 5) -> tuple[int, dict[str, int]] | None:
    for row_index in range(1, min(worksheet.max_row, search_rows) + 1):
        headers: dict[str, int] = {}
        for column_index in range(1, worksheet.max_column + 1):
            value = worksheet.cell(row_index, column_index).value
            if isinstance(value, str) and value.strip():
                headers[value.strip()] = column_index
        if required_headers.issubset(headers):
            return row_index, headers
    return None


def _normalize_store_key(value: str) -> str:
    normalized = re.sub(r"[^0-9A-Za-z가-힣]", "", value).lower()
    return normalized.strip()


def _find_card_sheet(path: Path):
    workbook = load_workbook(path, data_only=True)
    required_headers = {"가맹점명", "구분", "승인일자", "카드사명", "승인금액"}

    for worksheet in workbook.worksheets:
        extracted = _extract_header_row(worksheet, required_headers, search_rows=5)
        if extracted is not None:
            row_index, headers = extracted
            return worksheet, headers, row_index

    raise ParsingError(f"카드매출 파일에서 필수 헤더를 찾지 못했습니다: {path.name}")


def parse_card_sales(path: Path, config: AppConfig, allow_missing_date: bool = False) -> ParsedCardSales:
    worksheet, headers, header_row_index = _find_card_sheet(path)
    payment_by_source = config.payment_by_source_card_name
    by_store: dict[str, dict[str, float]] = {}
    account_date: str | None = None

    for row_index in range(header_row_index + 1, worksheet.max_row + 1):
        store_name = worksheet.cell(row_index, headers["가맹점명"]).value
        status = worksheet.cell(row_index, headers["구분"]).value
        card_name = worksheet.cell(row_index, headers["카드사명"]).value
        amount = worksheet.cell(row_index, headers["승인금액"]).value
        approval_date = worksheet.cell(row_index, headers["승인일자"]).value

        if not store_name or not card_name or not isinstance(amount, (int, float)):
            continue

        if account_date is None:
            account_date = _as_yyyymmdd(approval_date)

        payment = payment_by_source.get(str(card_name))
        if payment is None:
            continue

        # 카드매출은 승인금액만 사용한다.
        if status != "승인":
            continue

        store_bucket = by_store.setdefault(str(store_name), {})
        store_bucket[payment.key] = store_bucket.get(payment.key, 0.0) + float(amount)

    if account_date is None and not allow_missing_date:
        raise ParsingError(f"카드매출 파일에서 회계일을 찾지 못했습니다: {path.name}")

    return ParsedCardSales(account_date=account_date, by_store=by_store)


def _parse_daily_sales_new_format(path: Path, worksheet, allow_missing_date: bool) -> ParsedDailySales:
    required_headers = {"매장명", "영업일자", "총매출", "할인", "현금매출"}
    extracted = _extract_header_row(worksheet, required_headers, search_rows=5)
    if extracted is None:
        raise ParsingError(f"새양식 매출내역 헤더를 찾지 못했습니다: {path.name}")
    header_row_index, headers = extracted

    account_date = _extract_account_date_from_filename(path.name)
    if account_date is None:
        for row_index in range(header_row_index + 1, worksheet.max_row + 1):
            row_account_date = _as_yyyymmdd(worksheet.cell(row_index, headers["영업일자"]).value)
            if row_account_date:
                account_date = row_account_date
                break

    if account_date is None and not allow_missing_date:
        raise ParsingError(f"새양식 파일에서 회계일을 찾지 못했습니다: {path.name}")

    by_store: dict[str, DailySalesRow] = {}
    for row_index in range(header_row_index + 1, worksheet.max_row + 1):
        source_name = worksheet.cell(row_index, headers["매장명"]).value
        gross_sales = worksheet.cell(row_index, headers["총매출"]).value
        discount_amount = worksheet.cell(row_index, headers["할인"]).value
        cash_sales = worksheet.cell(row_index, headers["현금매출"]).value
        row_account_date = _as_yyyymmdd(worksheet.cell(row_index, headers["영업일자"]).value)

        if not isinstance(source_name, str) or not source_name.strip():
            continue
        if not isinstance(gross_sales, (int, float)):
            continue
        if account_date is not None and row_account_date is not None and row_account_date != account_date:
            continue

        key = source_name.strip()
        current = by_store.get(key)
        if current is None:
            by_store[key] = DailySalesRow(
                source_name=key,
                gross_sales=float(gross_sales or 0.0),
                discount_amount=float(discount_amount or 0.0),
                cash_sales=float(cash_sales or 0.0),
                electronic_money_sales=0.0,
            )
            continue

        by_store[key] = DailySalesRow(
            source_name=key,
            gross_sales=current.gross_sales + float(gross_sales or 0.0),
            discount_amount=current.discount_amount + float(discount_amount or 0.0),
            cash_sales=current.cash_sales + float(cash_sales or 0.0),
            electronic_money_sales=current.electronic_money_sales,
        )

    return ParsedDailySales(account_date=account_date, by_store=by_store)


def _parse_daily_sales_legacy_format(path: Path, worksheet, allow_missing_date: bool) -> ParsedDailySales:
    account_date = _extract_account_date_from_filename(path.name)
    if account_date is None and not allow_missing_date:
        raise ParsingError(f"당일 매출내역 파일명에서 회계일을 찾지 못했습니다: {path.name}")

    by_store: dict[str, DailySalesRow] = {}
    for row_index in range(3, worksheet.max_row + 1):
        source_name = worksheet.cell(row_index, 3).value
        gross_sales = worksheet.cell(row_index, 5).value
        discount_amount = worksheet.cell(row_index, 8).value
        cash_sales = worksheet.cell(row_index, 11).value
        electronic_money_sales = worksheet.cell(row_index, 20).value

        if not source_name or not isinstance(gross_sales, (int, float)):
            continue

        by_store[str(source_name)] = DailySalesRow(
            source_name=str(source_name),
            gross_sales=float(gross_sales or 0.0),
            discount_amount=float(discount_amount or 0.0),
            cash_sales=float(cash_sales or 0.0),
            electronic_money_sales=float(electronic_money_sales or 0.0),
        )

    return ParsedDailySales(account_date=account_date, by_store=by_store)


def parse_daily_sales(path: Path, allow_missing_date: bool = False) -> ParsedDailySales:
    workbook = load_workbook(path, data_only=True)
    worksheet = workbook.worksheets[0]

    new_format_headers = {"매장명", "영업일자", "총매출", "할인", "현금매출"}
    if _extract_header_row(worksheet, new_format_headers, search_rows=5) is not None:
        return _parse_daily_sales_new_format(path, worksheet, allow_missing_date)
    return _parse_daily_sales_legacy_format(path, worksheet, allow_missing_date)


def parse_settlement_orders(
    path: Path,
    *,
    source_names: set[str],
    store_aliases: dict[str, tuple[str, ...]] | None = None,
    expected_account_date: str | None = None,
    allow_missing_date: bool = False,
) -> ParsedSettlementOrders:
    workbook = load_workbook(path, data_only=True)
    worksheet = workbook.worksheets[0]
    required_headers = {"가맹점", "결제금액"}
    extracted = _extract_header_row(worksheet, required_headers, search_rows=10)
    if extracted is None:
        raise ParsingError(f"정산주문목록 파일에서 필수 헤더를 찾지 못했습니다: {path.name}")
    header_row_index, headers = extracted

    alias_index: dict[str, str] = {}
    for source_name in source_names:
        source_norm = _normalize_store_key(source_name)
        alias_index[source_norm] = source_name
        if store_aliases and source_name in store_aliases:
            for alias in store_aliases[source_name]:
                alias_norm = _normalize_store_key(alias)
                if alias_norm:
                    alias_index[alias_norm] = source_name

    account_date = expected_account_date or _extract_account_date_from_filename(path.name)
    merchant_totals: dict[str, float] = {}
    matched_merchants: dict[str, str] = {}
    by_store: dict[str, float] = {}
    unmatched_merchants: set[str] = set()

    for row_index in range(header_row_index + 1, worksheet.max_row + 1):
        merchant = worksheet.cell(row_index, headers["가맹점"]).value
        amount = worksheet.cell(row_index, headers["결제금액"]).value
        approval_date = (
            worksheet.cell(row_index, headers["결제 승인일"]).value
            if "결제 승인일" in headers
            else None
        )
        row_account_date = _as_yyyymmdd(approval_date)

        if account_date is None and row_account_date:
            account_date = row_account_date
        if account_date is not None and row_account_date is not None and row_account_date != account_date:
            continue

        if not isinstance(merchant, str) or not merchant.strip():
            continue
        if not isinstance(amount, (int, float)):
            continue

        merchant_name = merchant.strip()
        merchant_totals[merchant_name] = merchant_totals.get(merchant_name, 0.0) + float(amount)

    if account_date is None and not allow_missing_date:
        raise ParsingError(f"정산주문목록 파일에서 회계일을 찾지 못했습니다: {path.name}")

    for merchant_name, total_amount in merchant_totals.items():
        merchant_norm = _normalize_store_key(merchant_name)
        source_name = alias_index.get(merchant_norm)

        if source_name is None:
            # 보조 매칭: source_name 일부 문자열이 merchant 이름에 포함되거나 그 반대
            candidates: list[tuple[int, str]] = []
            for source in source_names:
                source_norm = _normalize_store_key(source)
                if not source_norm:
                    continue
                if source_norm in merchant_norm or merchant_norm in source_norm:
                    candidates.append((len(source_norm), source))
            if candidates:
                source_name = sorted(candidates, reverse=True)[0][1]

        if source_name is None:
            unmatched_merchants.add(merchant_name)
            continue

        matched_merchants[merchant_name] = source_name
        by_store[source_name] = by_store.get(source_name, 0.0) + total_amount

    return ParsedSettlementOrders(
        account_date=account_date,
        by_store=by_store,
        by_merchant=merchant_totals,
        matched_merchants=matched_merchants,
        unmatched_merchants=tuple(sorted(unmatched_merchants)),
    )


def _normalize_header_text(value: Any) -> str:
    if value is None:
        return ""
    if not isinstance(value, str):
        return ""
    return value.replace(" ", "").strip().lower()


def _coerce_amount(value: Any) -> float:
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        number = float(value)
        if math.isnan(number):
            return 0.0
        return number
    if not isinstance(value, str):
        return 0.0

    text = value.strip().replace(",", "")
    if not text:
        return 0.0
    if text.lower().startswith("0x"):
        return 0.0
    try:
        return float(text)
    except ValueError:
        return 0.0


def _as_yyyymmdd_from_serial(value: Any) -> str | None:
    if not isinstance(value, (int, float)):
        return None

    number = float(value)
    if math.isnan(number):
        return None

    # 20260329처럼 숫자 자체가 날짜인 경우도 수용한다.
    digits = str(int(number)) if number.is_integer() else ""
    if digits:
        normalized_digits = _normalize_date_digits(digits)
        if normalized_digits:
            return normalized_digits

    if number <= 0:
        return None

    try:
        converted = datetime(1899, 12, 30) + timedelta(days=number)
    except OverflowError:
        return None
    return converted.strftime("%Y%m%d")


def _resolve_sales_format_sheet_name(sheet_names: list[str], store: StoreMapping) -> str | None:
    source_no_station = store.source_name.replace("서울역점", "").strip()
    candidates = [
        store.output_name,
        store.source_name,
        source_no_station,
        store.output_name.replace(" ", ""),
        store.source_name.replace(" ", ""),
    ]
    for candidate in candidates:
        if candidate and candidate in sheet_names:
            return candidate

    for sheet_name in sheet_names:
        if (
            store.output_name in sheet_name
            or sheet_name in store.output_name
            or store.source_name in sheet_name
            or sheet_name in store.source_name
        ):
            return sheet_name
    return None


def _find_sales_format_header(rows: list[list[Any]]) -> tuple[int, dict[str, int]] | None:
    scan_limit = min(len(rows), 30)
    for row_index in range(scan_limit):
        row = rows[row_index]
        normalized_headers = [_normalize_header_text(cell_value) for cell_value in row]
        has_date_like = any("일" in header and "자" in header and "입금" not in header for header in normalized_headers)
        if not has_date_like:
            continue

        columns: dict[str, int] = {}
        gross_candidates: list[tuple[int, int]] = []
        card_candidates: list[tuple[int, int]] = []

        for column_index, header in enumerate(normalized_headers):
            if not header:
                continue

            if "일" in header and "자" in header and "입금" not in header:
                columns.setdefault("date", column_index)

            if "총매출액" in header:
                gross_candidates.append((0, column_index))
            elif header == "총매출":
                gross_candidates.append((1, column_index))
            elif "총매출" in header and "원가" not in header:
                gross_candidates.append((2, column_index))
            elif header == "매출액":
                gross_candidates.append((3, column_index))
            elif "순매출" in header:
                gross_candidates.append((4, column_index))

            if "할인" in header:
                columns.setdefault("discount", column_index)

            if header in {"현금", "총현금", "일반현금"} or ("현금" in header and "입금" not in header):
                columns.setdefault("cash", column_index)

            if "카드" in header and "카드사" not in header:
                if header == "총카드":
                    priority = 0
                elif header == "카드":
                    priority = 1
                elif header == "카드외":
                    priority = 2
                elif header == "일반카드":
                    priority = 3
                else:
                    priority = 9
                card_candidates.append((priority, column_index))

        if gross_candidates:
            columns["gross"] = sorted(gross_candidates)[0][1]
        if card_candidates:
            columns["card"] = sorted(card_candidates)[0][1]

        if "date" in columns and "gross" in columns:
            return row_index, columns
    return None


def _extract_sales_rows_from_sheet(
    *,
    rows: list[list[Any]],
    source_name: str,
    partner_code: str,
    date_from: str | None,
    date_to: str | None,
) -> tuple[list[SalesFormatInputRow], str | None]:
    header = _find_sales_format_header(rows)
    if header is None:
        return [], "시트에서 일자/매출 헤더를 찾지 못해 자동 입력 대상에서 제외했습니다."

    header_row_index, columns = header
    aggregated_by_date: dict[str, tuple[float, float, float]] = {}
    used_fallback_card = False

    for row in rows[header_row_index + 1 :]:
        date_cell = row[columns["date"]] if columns["date"] < len(row) else None
        account_date = _as_yyyymmdd(date_cell) or _as_yyyymmdd_from_serial(date_cell)
        if account_date is None:
            continue

        if date_from and account_date < date_from:
            continue
        if date_to and account_date > date_to:
            continue

        gross_cell = row[columns["gross"]] if columns["gross"] < len(row) else None
        gross_amount = _coerce_amount(gross_cell)

        discount_amount = 0.0
        if "discount" in columns:
            discount_cell = row[columns["discount"]] if columns["discount"] < len(row) else None
            discount_amount = abs(_coerce_amount(discount_cell))

        cash_amount = 0.0
        if "cash" in columns:
            cash_cell = row[columns["cash"]] if columns["cash"] < len(row) else None
            cash_amount = _coerce_amount(cash_cell)

        card_amount = 0.0
        if "card" in columns:
            card_cell = row[columns["card"]] if columns["card"] < len(row) else None
            card_amount = _coerce_amount(card_cell)

        if "card" not in columns or abs(card_amount) < 1e-9:
            card_amount = max(0.0, gross_amount - cash_amount - discount_amount)
            used_fallback_card = True

        if (
            abs(gross_amount) < 1e-9
            and abs(discount_amount) < 1e-9
            and abs(cash_amount) < 1e-9
            and abs(card_amount) < 1e-9
        ):
            continue

        current = aggregated_by_date.get(account_date)
        if current is None:
            aggregated_by_date[account_date] = (card_amount, cash_amount, discount_amount)
            continue

        aggregated_by_date[account_date] = (
            current[0] + card_amount,
            current[1] + cash_amount,
            current[2] + discount_amount,
        )

    parsed_rows = [
        SalesFormatInputRow(
            account_date=account_date,
            source_name=source_name,
            partner_code=partner_code,
            card_amount=amounts[0],
            cash_amount=amounts[1],
            discount_amount=amounts[2],
        )
        for account_date, amounts in sorted(aggregated_by_date.items())
    ]

    note = None
    if used_fallback_card:
        note = "카드 컬럼이 없거나 0인 구간은 총매출-현금-할인으로 카드금액을 보정했습니다."
    return parsed_rows, note


def parse_sales_format_input(
    path: Path,
    *,
    business: BusinessConfig,
    date_from: str | None = None,
    date_to: str | None = None,
) -> ParsedSalesFormatInput:
    if open_workbook is None:
        raise ParsingError("pyxlsb 패키지가 없어 .xlsb 파일을 읽을 수 없습니다. requirements 설치를 확인해 주세요.")

    if date_from and date_to and date_from > date_to:
        raise ParsingError("날짜 범위가 올바르지 않습니다. 시작일은 종료일보다 이전이어야 합니다.")
    if path.suffix.lower() != ".xlsb":
        raise ParsingError(f"매출 양식 파일은 .xlsb 형식이어야 합니다: {path.name}")

    rows: list[SalesFormatInputRow] = []
    parsed_store_names: set[str] = set()
    skipped_store_names: list[str] = []
    notes: list[str] = []

    with open_workbook(str(path)) as workbook:
        sheet_names = [str(sheet_name) for sheet_name in workbook.sheets]
        for store in business.active_stores:
            sheet_name = _resolve_sales_format_sheet_name(sheet_names, store)
            if sheet_name is None:
                skipped_store_names.append(store.source_name)
                continue

            with workbook.get_sheet(sheet_name) as worksheet:
                sheet_rows = [[cell.v for cell in row] for row in worksheet.rows()]

            extracted_rows, store_note = _extract_sales_rows_from_sheet(
                rows=sheet_rows,
                source_name=store.source_name,
                partner_code=store.partner_code,
                date_from=date_from,
                date_to=date_to,
            )

            if extracted_rows:
                parsed_store_names.add(store.source_name)
                rows.extend(extracted_rows)

            if store_note:
                notes.append(f"{store.output_name}: {store_note}")

    sorted_rows = sorted(rows, key=lambda item: (item.account_date, item.source_name))
    date_min = sorted_rows[0].account_date if sorted_rows else None
    date_max = sorted_rows[-1].account_date if sorted_rows else None

    return ParsedSalesFormatInput(
        rows=tuple(sorted_rows),
        parsed_store_count=len(parsed_store_names),
        skipped_store_names=tuple(sorted(skipped_store_names)),
        notes=tuple(notes),
        date_min=date_min,
        date_max=date_max,
    )


def parse_generated_voucher_settlement(
    path: Path,
    *,
    settlement_partner_code: str,
    settlement_management_data: str | None = None,
) -> ParsedVoucherSettlement:
    workbook = load_workbook(path, data_only=True)
    worksheet = workbook.worksheets[0]

    by_output_store: dict[str, float] = {}
    account_date: str | None = None
    row_index = 4
    note_pattern = re.compile(r"POS매출\((.+)\)")

    while True:
        docu_no = worksheet.cell(row_index, 1).value
        if docu_no is None:
            break

        actg_dt = worksheet.cell(row_index, 4).value
        if account_date is None and actg_dt:
            account_date = _as_yyyymmdd(str(actg_dt))

        account_code = worksheet.cell(row_index, 8).value
        amount = worksheet.cell(row_index, 9).value
        note_dc = worksheet.cell(row_index, 11).value
        partner_code = worksheet.cell(row_index, 12).value
        management_data = worksheet.cell(row_index, 15).value

        if str(account_code) != "102700":
            row_index += 1
            continue
        if str(partner_code) != settlement_partner_code:
            row_index += 1
            continue
        if settlement_management_data is not None and str(management_data) != settlement_management_data:
            row_index += 1
            continue
        if not isinstance(amount, (int, float)):
            row_index += 1
            continue
        if not isinstance(note_dc, str):
            row_index += 1
            continue

        matched = note_pattern.search(note_dc)
        if not matched:
            row_index += 1
            continue

        output_name = matched.group(1).strip()
        by_output_store[output_name] = by_output_store.get(output_name, 0.0) + float(amount)
        row_index += 1

    return ParsedVoucherSettlement(account_date=account_date, by_output_store=by_output_store)
