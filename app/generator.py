from __future__ import annotations

from copy import copy
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
import re
from tempfile import NamedTemporaryFile
from typing import Any

from openpyxl import load_workbook

from app.config import AppConfig, BusinessConfig, PaymentMethodConfig, StoreMapping
from app.parsers import (
    DailySalesRow,
    ParsedCardSales,
    ParsedDailySales,
    ParsedSettlementOrders,
    ParsedVoucherSettlement,
    ParsingError,
    parse_card_sales,
    parse_daily_sales,
    parse_generated_voucher_settlement,
    parse_settlement_orders,
)


@dataclass(frozen=True)
class VoucherLine:
    docu_no: str
    doline_sq: int
    row_line_sq: int
    actg_dt: str
    wrt_dt: str
    consul_dc: str | None
    drcrfg_cd: str
    acct_cd: str
    docu_amt: float
    docu_amt2: str
    note_dc: str
    partner_cd: str
    partner_nm: str
    cc_cd: str | None
    management_data: str | None
    row_no: str
    company_cd: str
    pc_cd: str
    gaap_cd: str
    docu_cd: str
    wrt_dept_cd: str
    wrt_emp_no: str
    insr_fg_cd: str

    def as_row(self) -> list[Any]:
        return [
            self.docu_no,
            self.doline_sq,
            self.row_line_sq,
            self.actg_dt,
            self.wrt_dt,
            self.consul_dc,
            self.drcrfg_cd,
            self.acct_cd,
            self.docu_amt,
            self.docu_amt2,
            self.note_dc,
            self.partner_cd,
            self.partner_nm,
            self.cc_cd,
            self.management_data,
            self.row_no,
            self.company_cd,
            self.pc_cd,
            self.gaap_cd,
            self.docu_cd,
            self.wrt_dept_cd,
            self.wrt_emp_no,
            self.insr_fg_cd,
        ]


@dataclass(frozen=True)
class GenerationArtifact:
    output_path: Path
    output_filename: str
    generated_store_names: tuple[str, ...]
    notes: tuple[str, ...]


@dataclass(frozen=True)
class SettlementVerificationArtifact:
    account_date: str
    settlement_store_count: int
    voucher_store_count: int
    matched_merchant_count: int
    unmatched_merchants: tuple[str, ...]
    differences: tuple[str, ...]


@dataclass(frozen=True)
class SalesTemplateFillArtifact:
    output_path: Path
    output_filename: str
    account_date: str
    filled_store_names: tuple[str, ...]
    skipped_store_names: tuple[str, ...]
    notes: tuple[str, ...]


SALES_TEMPLATE_PAYMENT_COLUMN_BY_KEY: dict[str, int] = {
    "hana": 13,
    "wechat": 14,
    "kb": 15,
    "nh": 16,
    "bc": 17,
    "hyundai": 18,
    "shinhan": 19,
    "samsung": 20,
    "lotte": 21,
}


VALIDATION_HEADER_COLUMNS: tuple[str, ...] = (
    "매장명",
    "현금매출액",
    "전자화폐",
    "카드매출액",
    "할인액",
    "총매출액",
    "입금기한",
    "입금일",
    "입금액",
    "잔액",
    "연체일",
    "연체료",
    "하나",
    "위쳇페이",
    "KB국민",
    "NH농협",
    "비씨",
    "현대",
    "신한",
    "삼성",
    "롯데",
    "합계",
    "카드매출파일 합계",
    "일치여부",
)


def save_upload_to_tempfile(filename: str, content: bytes) -> Path:
    suffix = Path(filename).suffix or ".xlsx"
    with NamedTemporaryFile(delete=False, suffix=suffix) as handle:
        handle.write(content)
        return Path(handle.name)


def _short_date(account_date: str) -> str:
    return account_date[2:]


def _format_amount_text(amount: float) -> str:
    rounded = round(amount)
    if abs(rounded - amount) < 1e-9:
        return str(int(rounded))
    return str(amount)


def _build_consul_dc(business: BusinessConfig, account_date: str) -> str:
    return business.consul_dc_format.format(
        date_short=_short_date(account_date),
        date_full=account_date,
        business=business.display_name,
    )


def _build_note_dc(business: BusinessConfig, account_date: str, output_name: str) -> str:
    return business.note_dc_format.format(
        date_short=_short_date(account_date),
        date_full=account_date,
        business=business.display_name,
        store=output_name,
    )


def _line_common(business: BusinessConfig, account_date: str) -> dict[str, str]:
    return {
        "docu_no": f"{business.doc_prefix}{account_date}",
        "actg_dt": account_date,
        "wrt_dt": account_date,
        "row_no": "1",
        "company_cd": business.company_code,
        "pc_cd": business.pc_code,
        "gaap_cd": business.gaap_code,
        "docu_cd": business.docu_code,
        "wrt_dept_cd": business.wrt_dept_code,
        "wrt_emp_no": business.wrt_emp_no,
        "insr_fg_cd": business.insr_fg_cd,
    }


def _payment_amounts_for_store(
    store: StoreMapping,
    card_sales: ParsedCardSales,
    daily_sales: ParsedDailySales,
    payment_methods: tuple[PaymentMethodConfig, ...],
) -> dict[str, float]:
    return dict(card_sales.by_store.get(store.source_name, {}))


def _build_store_lines(
    business: BusinessConfig,
    store: StoreMapping,
    daily_row: DailySalesRow,
    payment_amounts: dict[str, float],
    payment_methods: tuple[PaymentMethodConfig, ...],
    start_seq: int,
    account_date: str,
    include_consul_dc: bool,
    settlement_amount: float = 0.0,
    settlement_payment: PaymentMethodConfig | None = None,
    settlement_management_data: str | None = None,
) -> list[VoucherLine]:
    common = _line_common(business, account_date)
    note_dc = _build_note_dc(business, account_date, store.output_name)
    discount_amount = abs(daily_row.discount_amount or 0.0)
    receivable_total = sum(payment_amounts.values()) + settlement_amount
    debit_total = daily_row.cash_sales + receivable_total
    # 새 규칙:
    # - 할인(401511)은 할인액을 음수로
    # - 현금(103300) + 카드승인(102700)을 먼저 채운 뒤
    # - 남는 금액으로 401510을 계산해 차대합을 맞춘다.
    gross_sales_amount = debit_total + discount_amount
    lines: list[VoucherLine] = []
    seq = start_seq

    def add_line(
        *,
        drcrfg_cd: str,
        acct_cd: str,
        amount: float,
        partner_cd: str,
        partner_nm: str,
        cc_cd: str | None,
        management_data: str | None,
        consul_dc: str | None = None,
    ) -> None:
        nonlocal seq
        lines.append(
            VoucherLine(
                doline_sq=seq,
                row_line_sq=seq,
                consul_dc=consul_dc,
                drcrfg_cd=drcrfg_cd,
                acct_cd=acct_cd,
                docu_amt=amount,
                docu_amt2=_format_amount_text(amount),
                note_dc=note_dc,
                partner_cd=partner_cd,
                partner_nm=partner_nm,
                cc_cd=cc_cd,
                management_data=management_data,
                **common,
            )
        )
        seq += 1

    add_line(
        drcrfg_cd="2",
        acct_cd=business.account_codes["gross_sales"],
        amount=gross_sales_amount,
        partner_cd=store.partner_code,
        partner_nm=store.partner_name,
        cc_cd="1300",
        management_data=None,
        consul_dc=_build_consul_dc(business, account_date) if include_consul_dc else None,
    )

    if discount_amount:
        add_line(
            drcrfg_cd="2",
            acct_cd=business.account_codes["discount"],
            amount=(-1) * discount_amount,
            partner_cd=store.partner_code,
            partner_nm=store.partner_name,
            cc_cd="1300",
            management_data=None,
        )

    if daily_row.cash_sales:
        add_line(
            drcrfg_cd="1",
            acct_cd=business.account_codes["cash"],
            amount=daily_row.cash_sales,
            partner_cd=store.partner_code,
            partner_nm=store.partner_name,
            cc_cd=None,
            management_data=None,
        )

    for payment in payment_methods:
        amount = payment_amounts.get(payment.key, 0.0)
        if not amount:
            continue
        add_line(
            drcrfg_cd="1",
            acct_cd=business.account_codes["receivable"],
            amount=amount,
            partner_cd=payment.partner_code,
            partner_nm=payment.partner_name,
            cc_cd=None,
            management_data=payment.management_data,
        )

    if settlement_amount and settlement_payment is not None:
        add_line(
            drcrfg_cd="1",
            acct_cd=business.account_codes["receivable"],
            amount=settlement_amount,
            partner_cd=settlement_payment.partner_code,
            partner_nm=settlement_payment.partner_name,
            cc_cd=None,
            management_data=settlement_management_data or settlement_payment.management_data,
        )

    return lines


def _write_output(template_path: Path, output_path: Path, lines: list[VoucherLine]) -> None:
    workbook = load_workbook(template_path)
    worksheet = workbook.worksheets[0]

    if worksheet.max_row >= 4:
        worksheet.delete_rows(4, worksheet.max_row - 3)

    for row_index, line in enumerate(lines, start=4):
        for column_index, value in enumerate(line.as_row(), start=1):
            worksheet.cell(row_index, column_index).value = value

    workbook.save(output_path)


def _normalize_store_key(value: str) -> str:
    return re.sub(r"[^0-9A-Za-z가-힣]", "", value).lower().strip()


def _as_yyyymmdd(value: Any) -> str | None:
    if isinstance(value, datetime):
        return value.strftime("%Y%m%d")
    if isinstance(value, date):
        return value.strftime("%Y%m%d")
    if isinstance(value, str):
        digits = re.sub(r"[^0-9]", "", value)
        if len(digits) == 6:
            digits = f"20{digits}"
        if len(digits) == 8:
            try:
                datetime.strptime(digits, "%Y%m%d")
                return digits
            except ValueError:
                return None
    return None


def _resolve_account_date(
    *,
    parsed_card: ParsedCardSales,
    parsed_daily: ParsedDailySales,
    manual_account_date: str | None,
    notes: list[str],
) -> str:
    if manual_account_date is not None:
        if parsed_card.account_date and parsed_card.account_date != manual_account_date:
            notes.append(
                f"카드 파일 회계일({parsed_card.account_date}) 대신 입력값({manual_account_date})을 사용했습니다."
            )
        if parsed_daily.account_date and parsed_daily.account_date != manual_account_date:
            notes.append(
                f"당일 파일 회계일({parsed_daily.account_date}) 대신 입력값({manual_account_date})을 사용했습니다."
            )
        return manual_account_date

    if parsed_card.account_date != parsed_daily.account_date:
        raise ParsingError(
            "카드매출/당일매출 파일의 회계일이 다릅니다. "
            f"카드={parsed_card.account_date}, 당일={parsed_daily.account_date}"
        )
    if parsed_card.account_date is None:
        raise ParsingError("회계일자를 확인할 수 없습니다. 회계일자를 직접 입력해 주세요.")
    return parsed_card.account_date


def _build_store_candidates(store: StoreMapping) -> list[str]:
    candidates = [
        store.output_name,
        store.source_name,
        store.source_name.replace("서울역점", "").strip(),
        store.output_name.replace("서울역점", "").strip(),
    ]
    unique: list[str] = []
    for candidate in candidates:
        if candidate and candidate not in unique:
            unique.append(candidate)
    return unique


def _match_name_from_keys(
    *,
    keys: list[str],
    store: StoreMapping,
) -> str | None:
    candidates = _build_store_candidates(store)
    for candidate in candidates:
        if candidate in keys:
            return candidate

    normalized_index = {key: _normalize_store_key(key) for key in keys}
    candidate_norms = [norm for norm in (_normalize_store_key(item) for item in candidates) if norm]

    for candidate_norm in candidate_norms:
        for original, normalized in normalized_index.items():
            if normalized == candidate_norm:
                return original

    scored: list[tuple[int, str]] = []
    for candidate_norm in candidate_norms:
        for original, normalized in normalized_index.items():
            if not normalized:
                continue
            if candidate_norm in normalized or normalized in candidate_norm:
                scored.append((len(normalized), original))
    if scored:
        return sorted(scored, reverse=True)[0][1]
    return None


def _find_template_row_for_account_date(worksheet, account_date: str) -> int | None:
    for row_index in range(4, 35):
        row_date = _as_yyyymmdd(worksheet.cell(row_index, 1).value)
        if row_date == account_date:
            return row_index
    return None


def _sales_template_output_filename(template_file_path: Path, account_date: str) -> str:
    return f"{template_file_path.stem}_{account_date}_자동입력.xlsx"


def _copy_validation_template_sheet(template_path: Path, workbook, sheet_title: str):
    template_workbook = load_workbook(template_path)
    template_sheet = template_workbook.worksheets[0]
    worksheet = workbook.create_sheet(title=sheet_title, index=0)

    worksheet.sheet_format = copy(template_sheet.sheet_format)
    worksheet.sheet_properties = copy(template_sheet.sheet_properties)
    worksheet.page_margins = copy(template_sheet.page_margins)
    worksheet.page_setup = copy(template_sheet.page_setup)
    worksheet.print_options = copy(template_sheet.print_options)
    worksheet.freeze_panes = template_sheet.freeze_panes

    for column_key, column_dimension in template_sheet.column_dimensions.items():
        target_dimension = worksheet.column_dimensions[column_key]
        target_dimension.width = column_dimension.width
        target_dimension.hidden = column_dimension.hidden
        target_dimension.outlineLevel = column_dimension.outlineLevel
        target_dimension.collapsed = column_dimension.collapsed
        target_dimension.bestFit = column_dimension.bestFit

    for row_index, row_dimension in template_sheet.row_dimensions.items():
        target_dimension = worksheet.row_dimensions[row_index]
        target_dimension.height = row_dimension.height
        target_dimension.hidden = row_dimension.hidden
        target_dimension.outlineLevel = row_dimension.outlineLevel
        target_dimension.collapsed = row_dimension.collapsed

    for merged_range in template_sheet.merged_cells.ranges:
        worksheet.merge_cells(str(merged_range))

    for row in template_sheet.iter_rows(
        min_row=1,
        max_row=template_sheet.max_row,
        min_col=1,
        max_col=template_sheet.max_column,
    ):
        for source_cell in row:
            target_cell = worksheet.cell(
                row=source_cell.row,
                column=source_cell.column,
                value=source_cell.value,
            )
            if source_cell.has_style:
                target_cell._style = copy(source_cell._style)
            if source_cell.hyperlink is not None:
                target_cell._hyperlink = copy(source_cell.hyperlink)
            if source_cell.comment is not None:
                target_cell.comment = copy(source_cell.comment)

    return worksheet


def _create_validation_sheet(
    *,
    workbook,
    account_date: str,
    validation_rows: list[dict[str, Any]],
    validation_template_file_path: Path | None,
) -> None:
    sheet_title = f"({account_date}) 검증"
    if sheet_title in workbook.sheetnames:
        del workbook[sheet_title]

    template_exists = validation_template_file_path is not None and validation_template_file_path.exists()
    if template_exists:
        worksheet = _copy_validation_template_sheet(validation_template_file_path, workbook, sheet_title)
    else:
        worksheet = workbook.create_sheet(title=sheet_title, index=0)

    worksheet.cell(1, 1).value = "일자"
    worksheet.cell(1, 2).value = datetime.strptime(account_date, "%Y%m%d").date()

    for column_index, header in enumerate(VALIDATION_HEADER_COLUMNS, start=1):
        if worksheet.cell(2, column_index).value is None:
            worksheet.cell(2, column_index).value = header

    if worksheet.max_row >= 3:
        worksheet.delete_rows(3, worksheet.max_row - 2)

    for offset, row in enumerate(validation_rows):
        row_index = 3 + offset
        payment_amounts: dict[str, float] = row["payment_amounts"]

        worksheet.cell(row_index, 1).value = row["store_name"]
        worksheet.cell(row_index, 2).value = float(row["cash_sales"])
        worksheet.cell(row_index, 3).value = float(row["electronic_money_sales"])
        worksheet.cell(row_index, 4).value = float(row["card_sales"])
        worksheet.cell(row_index, 5).value = float(row["discount_amount"])
        worksheet.cell(row_index, 6).value = float(row["gross_sales"])

        for payment_key, column in SALES_TEMPLATE_PAYMENT_COLUMN_BY_KEY.items():
            worksheet.cell(row_index, column).value = float(payment_amounts.get(payment_key, 0.0) or 0.0)

        worksheet.cell(row_index, 22).value = f"=SUM(M{row_index}:U{row_index})"
        worksheet.cell(row_index, 23).value = float(row["card_sales"])
        worksheet.cell(row_index, 24).value = f"=V{row_index}=W{row_index}"


def generate_voucher(
    *,
    business: BusinessConfig,
    card_file_path: Path,
    daily_file_path: Path,
    template_file_path: Path,
    config: AppConfig,
    output_dir: Path,
    manual_account_date: str | None = None,
    settlement_order_file_path: Path | None = None,
    settlement_store_aliases: dict[str, tuple[str, ...]] | None = None,
) -> GenerationArtifact:
    notes: list[str] = []
    allow_missing_date = manual_account_date is not None
    parsed_card = parse_card_sales(card_file_path, config, allow_missing_date=allow_missing_date)
    parsed_daily = parse_daily_sales(daily_file_path, allow_missing_date=allow_missing_date)
    account_date = _resolve_account_date(
        parsed_card=parsed_card,
        parsed_daily=parsed_daily,
        manual_account_date=manual_account_date,
        notes=notes,
    )

    parsed_settlement: ParsedSettlementOrders | None = None
    settlement_payment = config.payment_by_key.get(config.settlement_partner_payment_key)
    if settlement_order_file_path is not None:
        if settlement_payment is None:
            raise ParsingError(
                f"정산주문목록용 결제수단 키({config.settlement_partner_payment_key})를 payment_methods에서 찾지 못했습니다."
            )
        parsed_settlement = parse_settlement_orders(
            settlement_order_file_path,
            source_names=business.source_store_names,
            store_aliases=settlement_store_aliases,
            expected_account_date=account_date,
            allow_missing_date=True,
        )
        notes.append(
            "정산주문목록 매칭: "
            f"{len(parsed_settlement.matched_merchants)}개 가맹점, "
            f"{len(parsed_settlement.unmatched_merchants)}개 미매칭"
        )
        if parsed_settlement.unmatched_merchants:
            notes.append("정산주문목록 미매칭 가맹점: " + ", ".join(parsed_settlement.unmatched_merchants))

    generated_store_names: list[str] = []
    lines: list[VoucherLine] = []
    next_seq = 1
    first_line = True

    for store in business.active_stores:
        daily_row = parsed_daily.by_store.get(store.source_name)
        if not daily_row:
            continue

        amounts = _payment_amounts_for_store(
            store=store,
            card_sales=parsed_card,
            daily_sales=parsed_daily,
            payment_methods=config.payment_methods_in_order,
        )
        settlement_amount = parsed_settlement.by_store.get(store.source_name, 0.0) if parsed_settlement else 0.0

        has_amount = any(
            [
                daily_row.gross_sales,
                daily_row.discount_amount,
                daily_row.cash_sales,
                daily_row.electronic_money_sales,
                sum(amounts.values()),
                settlement_amount,
            ]
        )
        if not has_amount or daily_row.gross_sales <= 0:
            continue

        store_lines = _build_store_lines(
            business=business,
            store=store,
            daily_row=daily_row,
            payment_amounts=amounts,
            payment_methods=config.payment_methods_in_order,
            start_seq=next_seq,
            account_date=account_date,
            include_consul_dc=first_line,
            settlement_amount=settlement_amount,
            settlement_payment=settlement_payment,
            settlement_management_data=config.settlement_management_data,
        )
        lines.extend(store_lines)
        next_seq += len(store_lines)
        first_line = False
        generated_store_names.append(store.output_name)

        if store.manual_review_reason:
            notes.append(f"{store.output_name}: {store.manual_review_reason}")

    input_store_names = {
        name
        for name, row in parsed_daily.by_store.items()
        if row.gross_sales or row.discount_amount or row.cash_sales or row.electronic_money_sales
    }
    active_source_names = business.source_store_names
    intentionally_excluded_source_names = {
        store.source_name
        for store in business.stores
        if not store.enabled
    }
    skipped = sorted(input_store_names - active_source_names - intentionally_excluded_source_names)
    if skipped:
        notes.append("자동 생성 제외 매장(입력에는 존재): " + ", ".join(skipped))

    output_filename = f"{business.display_name}POS매장임대을매출_{account_date}.xlsx"
    output_path = output_dir / output_filename
    _write_output(template_file_path, output_path, lines)

    return GenerationArtifact(
        output_path=output_path,
        output_filename=output_filename,
        generated_store_names=tuple(generated_store_names),
        notes=tuple(notes),
    )


def generate_sales_template_auto_input(
    *,
    business: BusinessConfig,
    card_file_path: Path,
    daily_file_path: Path,
    sales_template_file_path: Path,
    config: AppConfig,
    output_dir: Path,
    manual_account_date: str | None = None,
    validation_template_file_path: Path | None = None,
    settlement_order_file_path: Path | None = None,
    settlement_store_aliases: dict[str, tuple[str, ...]] | None = None,
) -> SalesTemplateFillArtifact:
    output_dir.mkdir(parents=True, exist_ok=True)
    notes: list[str] = []
    allow_missing_date = manual_account_date is not None
    parsed_card = parse_card_sales(card_file_path, config, allow_missing_date=allow_missing_date)
    parsed_daily = parse_daily_sales(daily_file_path, allow_missing_date=allow_missing_date)
    account_date = _resolve_account_date(
        parsed_card=parsed_card,
        parsed_daily=parsed_daily,
        manual_account_date=manual_account_date,
        notes=notes,
    )
    parsed_settlement: ParsedSettlementOrders | None = None
    settlement_payment = config.payment_by_key.get(config.settlement_partner_payment_key)
    settlement_daily_field = settlement_payment.daily_field if settlement_payment else None
    if settlement_order_file_path is not None:
        if settlement_payment is None:
            raise ParsingError(
                f"정산주문목록용 결제수단 키({config.settlement_partner_payment_key})를 payment_methods에서 찾지 못했습니다."
            )
        parsed_settlement = parse_settlement_orders(
            settlement_order_file_path,
            source_names=business.source_store_names,
            store_aliases=settlement_store_aliases,
            expected_account_date=account_date,
            allow_missing_date=True,
        )
        notes.append(
            "정산주문목록 매칭: "
            f"{len(parsed_settlement.matched_merchants)}개 가맹점, "
            f"{len(parsed_settlement.unmatched_merchants)}개 미매칭"
        )
        if parsed_settlement.unmatched_merchants:
            notes.append("정산주문목록 미매칭 가맹점: " + ", ".join(parsed_settlement.unmatched_merchants))

    workbook = load_workbook(sales_template_file_path)
    sheet_names = workbook.sheetnames
    daily_keys = list(parsed_daily.by_store.keys())
    card_keys = list(parsed_card.by_store.keys())

    filled_store_names: list[str] = []
    skipped_store_names: list[str] = []
    validation_rows: list[dict[str, Any]] = []
    settlement_applied_count = 0
    settlement_unapplied_stores: list[str] = []

    for store in business.active_stores:
        sheet_name = _match_name_from_keys(keys=sheet_names, store=store)
        if sheet_name is None:
            skipped_store_names.append(store.output_name)
            continue

        daily_store_key = _match_name_from_keys(keys=daily_keys, store=store)
        if daily_store_key is None:
            skipped_store_names.append(store.output_name)
            continue
        daily_row = parsed_daily.by_store[daily_store_key]

        card_store_key = _match_name_from_keys(keys=card_keys, store=store)
        payment_amounts = parsed_card.by_store.get(card_store_key, {}) if card_store_key else {}
        settlement_amount = parsed_settlement.by_store.get(store.source_name, 0.0) if parsed_settlement else 0.0
        cash_sales = float(daily_row.cash_sales or 0.0)
        electronic_money_sales = float(daily_row.electronic_money_sales or 0.0)
        gross_sales = float(daily_row.gross_sales or 0.0)
        discount_amount = abs(float(daily_row.discount_amount or 0.0))
        settlement_applied = False
        settlement_unapplied = False

        if settlement_amount:
            normalized_settlement_field = (settlement_daily_field or "").replace(" ", "")
            if normalized_settlement_field.startswith("전자화폐"):
                settlement_value = float(settlement_amount)
                gross_sales = gross_sales - electronic_money_sales + settlement_value
                electronic_money_sales = settlement_value
                settlement_applied = True
            elif normalized_settlement_field.startswith("현금"):
                settlement_value = float(settlement_amount)
                gross_sales = gross_sales - cash_sales + settlement_value
                cash_sales = settlement_value
                settlement_applied = True
            else:
                settlement_unapplied = True

        worksheet = workbook[sheet_name]
        target_row = _find_template_row_for_account_date(worksheet, account_date)
        if target_row is None:
            skipped_store_names.append(store.output_name)
            continue

        if settlement_applied:
            settlement_applied_count += 1
        elif settlement_unapplied:
            settlement_unapplied_stores.append(store.output_name)

        # 입력영역(4~34행)만 수정. 36행 이후는 건드리지 않는다.
        worksheet.cell(target_row, 2).value = cash_sales
        worksheet.cell(target_row, 3).value = electronic_money_sales
        worksheet.cell(target_row, 5).value = discount_amount

        for column in range(13, 22):
            worksheet.cell(target_row, column).value = 0.0
        for payment in config.payment_methods_in_order:
            column = SALES_TEMPLATE_PAYMENT_COLUMN_BY_KEY.get(payment.key)
            if column is None:
                continue
            amount = float(payment_amounts.get(payment.key, 0.0) or 0.0)
            worksheet.cell(target_row, column).value = amount

        card_sales_amount = sum(float(value or 0.0) for value in payment_amounts.values())
        validation_rows.append(
            {
                "store_name": store.output_name,
                "cash_sales": cash_sales,
                "electronic_money_sales": electronic_money_sales,
                "discount_amount": discount_amount,
                "gross_sales": gross_sales,
                "card_sales": float(card_sales_amount),
                "payment_amounts": {
                    key: float(payment_amounts.get(key, 0.0) or 0.0)
                    for key in SALES_TEMPLATE_PAYMENT_COLUMN_BY_KEY
                },
            }
        )
        filled_store_names.append(store.output_name)

    if not filled_store_names:
        raise ParsingError("채워 넣을 매장이 없습니다. 시트명/매장명 매핑 또는 입력 파일 내용을 확인해 주세요.")

    output_filename = _sales_template_output_filename(sales_template_file_path, account_date)
    output_path = output_dir / output_filename
    _create_validation_sheet(
        workbook=workbook,
        account_date=account_date,
        validation_rows=validation_rows,
        validation_template_file_path=validation_template_file_path,
    )
    workbook.save(output_path)

    if skipped_store_names:
        notes.append("자동 입력 제외(시트/입력 매칭 실패): " + ", ".join(sorted(set(skipped_store_names))))
    if validation_template_file_path is None or not validation_template_file_path.exists():
        notes.append("검증시트 템플릿을 찾지 못해 기본 헤더로 검증 시트를 생성했습니다.")
    if parsed_settlement is not None:
        if settlement_applied_count:
            notes.append(
                f"정산 보정 반영: {settlement_applied_count}개 매장 "
                f"(반영 필드: {settlement_daily_field or '미설정'})"
            )
        if settlement_unapplied_stores:
            notes.append(
                "정산 보정 미반영 매장(결제수단 daily_field 확인 필요): "
                + ", ".join(sorted(set(settlement_unapplied_stores)))
            )

    return SalesTemplateFillArtifact(
        output_path=output_path,
        output_filename=output_filename,
        account_date=account_date,
        filled_store_names=tuple(filled_store_names),
        skipped_store_names=tuple(sorted(set(skipped_store_names))),
        notes=tuple(notes),
    )


def verify_settlement_against_voucher(
    *,
    business: BusinessConfig,
    voucher_file_path: Path,
    settlement_order_file_path: Path,
    settlement_partner_code: str,
    settlement_management_data: str | None = None,
    settlement_store_aliases: dict[str, tuple[str, ...]] | None = None,
    manual_account_date: str | None = None,
) -> SettlementVerificationArtifact:
    parsed_voucher: ParsedVoucherSettlement = parse_generated_voucher_settlement(
        voucher_file_path,
        settlement_partner_code=settlement_partner_code,
        settlement_management_data=settlement_management_data,
    )
    expected_account_date = manual_account_date or parsed_voucher.account_date
    if expected_account_date is None:
        raise ParsingError("검증 기준 회계일자를 찾지 못했습니다. 회계일자를 직접 입력해 주세요.")

    parsed_settlement: ParsedSettlementOrders = parse_settlement_orders(
        settlement_order_file_path,
        source_names=business.source_store_names,
        store_aliases=settlement_store_aliases,
        expected_account_date=expected_account_date,
        allow_missing_date=True,
    )

    source_to_output = {store.source_name: store.output_name for store in business.active_stores}
    expected_by_output: dict[str, float] = {}
    for source_name, amount in parsed_settlement.by_store.items():
        output_name = source_to_output.get(source_name)
        if output_name is None:
            continue
        expected_by_output[output_name] = expected_by_output.get(output_name, 0.0) + amount

    actual_by_output = parsed_voucher.by_output_store
    all_output_names = sorted(set(expected_by_output) | set(actual_by_output))
    differences: list[str] = []
    for output_name in all_output_names:
        expected_amount = expected_by_output.get(output_name, 0.0)
        actual_amount = actual_by_output.get(output_name, 0.0)
        diff = actual_amount - expected_amount
        if abs(diff) > 0.1:
            differences.append(
                f"{output_name}: 전표={_format_amount_text(actual_amount)}, "
                f"정산주문목록={_format_amount_text(expected_amount)}, 차이={_format_amount_text(diff)}"
            )

    return SettlementVerificationArtifact(
        account_date=expected_account_date,
        settlement_store_count=len(expected_by_output),
        voucher_store_count=len(actual_by_output),
        matched_merchant_count=len(parsed_settlement.matched_merchants),
        unmatched_merchants=parsed_settlement.unmatched_merchants,
        differences=tuple(differences),
    )

